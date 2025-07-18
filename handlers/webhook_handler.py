from handlers.image_handler import process_image_message
from handlers.text_handler import process_text_message
from handlers.pdf_handler import process_pdf_message
from handlers.csv_handler import (
    xlsx_with_summary_update,  # サマリ生成
    normalize_df,
    create_order_list_sheet,
    create_order_sheets,       # ← 注文書自動作成
    autofit_columns,
)
from handlers.file_handler import get_or_create_folder, drive_service
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from config import CSV_FORMAT_PATH, SHARED_DRIVE_ID, ORDER_SUMMARY_FOLDER_ID

import os
import pytz
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openai import OpenAI
import requests
from handlers.csv_handler import migrate_prev_day_sheets_to_today
from openpyxl.utils import get_column_letter
import unicodedata

def handle_webhook(request):
    data = request.get_json()
    events = data.get('events', [])
    if not events:
        return 'OK', 200

    event = events[0]
    message_type = event.get('message', {}).get('type')

    if message_type == 'text':
        user_text = event['message'].get('text', '').strip()
        JST = pytz.timezone('Asia/Tokyo')
        today = datetime.now(JST).strftime('%Y%m%d')

        # Driveの「受注集計＞{today}＞集計結果」までのIDを取得
        try:
            root_id = get_or_create_folder('受注集計')
            date_id = get_or_create_folder(today, parent_id=root_id)
            csv_folder_id = get_or_create_folder('集計結果', parent_id=date_id)
        except Exception as e:
            print(f"DriveフォルダID取得エラー: {e}")
            return 'OK', 200

        filename = f'集計結果_{today}.xlsx'
        file_path = f"/tmp/{filename}"

        # Drive内でファイルを検索
        query = f"name = '{filename}' and '{csv_folder_id}' in parents and trashed = false"
        response = drive_service.files().list(
            q=query,
            fields='files(id)',
            driveId=SHARED_DRIVE_ID,
            corpora='drive',
            includeItemsFromAllDrives=True,
            supportsAllDrives=True
        ).execute()
        files = response.get('files', [])
        if not files:
            print("集計ファイルが見つかりません")
            # --- 新規作成ロジックここから ---
            df_empty = pd.DataFrame(columns=["顧客", "発注者", "商品名", "サイズ", "数量", "単位", "納品希望日", "納品場所", "時間", "社内担当者", "備考"])
            df_empty.to_excel(file_path, index=False)
            media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file_metadata = {'name': filename, 'parents': [csv_folder_id]}
            drive_service.files().create(
                body=file_metadata, media_body=media, fields='id', supportsAllDrives=True
            ).execute()
            print("空の集計ファイルを新規作成しアップロードしました")
            response = drive_service.files().list(
                q=query,
                fields='files(id)',
                driveId=SHARED_DRIVE_ID,
                corpora='drive',
                includeItemsFromAllDrives=True,
                supportsAllDrives=True
            ).execute()
            files = response.get('files', [])
            if not files:
                print("空ファイル作成後も取得できません")
                return 'OK', 200

        file_id = files[0]['id']
        # ファイルを一時保存
        try:
            request_dl = drive_service.files().get_media(fileId=file_id)
            with open(file_path, 'wb') as fh:
                downloader = MediaIoBaseDownload(fh, request_dl)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
        except Exception as e:
            print(f"DriveファイルDLエラー: {e}")
            return 'OK', 200

        # =====================
        # サマリ生成
        # =====================
        if user_text == '集計サマリ作成':
            try:
                openai_client = OpenAI()
                today = datetime.now(JST).strftime('%Y%m%d')
                sheet_name = f'集計結果_{today}'
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                df_norm = normalize_df(df, openai_client)
                xlsx_with_summary_update(df_norm, file_path, openai_client)

                # 再アップロード
                media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                drive_service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()
                print(f"サマリ生成後にDriveへ再アップロード完了: {filename}")

            except Exception as e:
                print(f"サマリ生成またはDriveアップロードエラー: {e}")

            return 'OK', 200

        # =====================
        # ピッキングリスト作成
        # =====================
        if user_text == 'ピッキングリスト作成':
            try:
                df = pd.read_excel(file_path, sheet_name=None)
                main_sheet_name = f"集計結果_{today}"
                if main_sheet_name not in df:
                    # なければ1枚目 fallback（旧方式）なども可
                    print(f"{main_sheet_name} シートがありません。既存シートを利用します")
                    main_sheet_name = list(df.keys())[0]
                main_df = df[main_sheet_name]

                # 本日納品希望分だけ
                pick_df = main_df[main_df['納品希望日'].astype(str) == today]

                # --- 受注残(前日データ)も存在すれば追加 ---
                if '受注残(前日データ)' in df:
                    prev_juchu_df = df['受注残(前日データ)']
                    # カラム名が0行目になっていれば修正（Excel/Pandas由来でズレることあり）
                    if not isinstance(prev_juchu_df.columns[0], str):
                        prev_juchu_df.columns = prev_juchu_df.iloc[0]
                        prev_juchu_df = prev_juchu_df[1:]
                    prev_pick_df = prev_juchu_df[prev_juchu_df['納品希望日'].astype(str) == today]
                    # 本体とマージ
                    pick_df = pd.concat([pick_df, prev_pick_df], ignore_index=True)

                wb = load_workbook(file_path)
                # すでに存在すれば削除
                if 'ピッキングリスト' in wb.sheetnames:
                    ws = wb['ピッキングリスト']
                    wb.remove(ws)
                ws = wb.create_sheet('ピッキングリスト')
                ws.append(list(main_df.columns))  # 1行目
                for row in pick_df.itertuples(index=False, name=None):
                    ws.append(row)
                for ws in wb.worksheets:
                    autofit_columns(ws)
                wb.save(file_path)

                # 再アップロード
                media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                drive_service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()
                print("ピッキングリスト作成＆Drive再アップロード完了！")
            except Exception as e:
                print(f"ピッキングリスト作成またはDriveアップロードエラー: {e}")
            return 'OK', 200

        # =====================
        # 発注リスト作成
        # =====================
        if user_text == '発注リスト作成':
            try:
                tag_xlsx_path = f"/tmp/タグ付け表.xlsx"
                # 「親フォルダ直下の全ファイル」一括取得
                files = drive_service.files().list(
                    q=f"'{root_id}' in parents and trashed = false",
                    fields="files(id, name)",
                    driveId=SHARED_DRIVE_ID,
                    corpora='drive',
                    includeItemsFromAllDrives=True,
                    supportsAllDrives=True
                ).execute().get('files', [])
                
                # Python側で正規化比較
                target_name = unicodedata.normalize('NFC', 'タグ付け表.xlsx')
                tag_file_id = None
                for f in files:
                    f_name = unicodedata.normalize('NFC', f['name'])
                    if f_name == target_name:
                        tag_file_id = f['id']
                        break

                if not tag_file_id:
                    print("タグ付け表.xlsxが見つかりません")
                    return 'OK', 200

                tag_dl = drive_service.files().get_media(fileId=tag_file_id)
                with open(tag_xlsx_path, 'wb') as ftag:
                    downloader = MediaIoBaseDownload(ftag, tag_dl)
                    done = False
                    while not done:
                        status, done = downloader.next_chunk()

                # シート作成
                ok = create_order_list_sheet(file_path, tag_xlsx_path)
                if not ok:
                    print("注文リストシート作成に失敗")
                    return 'OK', 200

                # Drive再アップロード
                media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                drive_service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()
                print("注文リスト作成＆Drive再アップロード完了！")

            except Exception as e:
                print(f"注文リスト作成またはDriveアップロードエラー: {e}")
            return 'OK', 200

        # =====================
        # 発注書作成（←ここでcsv_handlerからインポートした関数を使用）
        # =====================
        if user_text == '注文書作成':
            try:
                ok = create_order_sheets(date_id, csv_folder_id, today, drive_service)
                if not ok:
                    print("注文書作成に失敗")
                    return 'OK', 200
                print("注文書自動作成完了！")
            except Exception as e:
                print(f"注文書作成エラー: {e}")
            return 'OK', 200

        # =====================
        # 受注残＋発注残シート同時作成
        # =====================
        if user_text == '受注残と発注残の作成':
            try:
                wb = load_workbook(file_path)
                JST = pytz.timezone('Asia/Tokyo')
                tomorrow = (datetime.now(JST) + timedelta(days=1)).strftime('%Y%m%d')

                # --- 受注残シート ---
                main_sheet_name = f"集計結果_{today}"
                main_df = pd.DataFrame(wb[main_sheet_name].values)
                main_df.columns = main_df.iloc[0]
                main_df = main_df[1:]
                main_df = main_df.loc[:, main_df.columns.notna() & (main_df.columns != "None")]
                main_df.columns = main_df.columns.map(lambda x: str(x).strip())
                main_df = main_df.loc[:, ~main_df.columns.duplicated()]
                remaining_df = main_df[main_df['納品希望日'].astype(str) >= tomorrow]

                # 受注残(前日データ)からの追加
                if '受注残(前日データ)' in wb.sheetnames:
                    prev_df = pd.DataFrame(wb['受注残(前日データ)'].values)
                    prev_df.columns = prev_df.iloc[0]
                    prev_df = prev_df[1:]
                    prev_df = prev_df.loc[:, prev_df.columns.notna() & (prev_df.columns != "None")]
                    prev_df.columns = prev_df.columns.map(lambda x: str(x).strip())
                    prev_df = prev_df.loc[:, ~prev_df.columns.duplicated()]
                    # カラム揃え
                    prev_df = prev_df.reindex(columns=remaining_df.columns, fill_value="")
                    prev_add_df = prev_df[prev_df['納品希望日'].astype(str) >= tomorrow]
                    # 合体
                    remaining_df = pd.concat([remaining_df, prev_add_df], ignore_index=True)
                    remaining_df = remaining_df.drop_duplicates()

                # 受注残シート作成
                if '受注残' in wb.sheetnames:
                    del wb['受注残']
                ws_juchu = wb.create_sheet('受注残')
                ws_juchu.append(list(remaining_df.columns))
                for row in remaining_df.itertuples(index=False, name=None):
                    if all([cell is None or str(cell).strip() == "" for cell in row]):
                        continue
                    ws_juchu.append(row)

                # --- 注文残シート ---
                # まず既存ロジックで作成
                from handlers.csv_handler import create_order_remains_sheet_from_wb
                ok = create_order_remains_sheet_from_wb(wb)
                if not ok:
                    print("発注残作成に失敗")
                else:
                    print("注文残シート作成成功")

                # 注文残(前日データ)からの追加
                if '注文残(前日データ)' in wb.sheetnames and '注文残' in wb.sheetnames:
                    order_zan_ws = wb['注文残']
                    # Pandasで加工
                    order_zan_df = pd.DataFrame(order_zan_ws.values)
                    order_zan_df.columns = order_zan_df.iloc[0]
                    order_zan_df = order_zan_df[1:]
                    prev_df = pd.DataFrame(wb['注文残(前日データ)'].values)
                    prev_df.columns = prev_df.iloc[0]
                    prev_df = prev_df[1:]
                    prev_df = prev_df.loc[:, prev_df.columns.notna() & (prev_df.columns != "None")]
                    prev_df.columns = prev_df.columns.map(lambda x: str(x).strip())
                    prev_df = prev_df.loc[:, ~prev_df.columns.duplicated()]
                    prev_df = prev_df.reindex(columns=order_zan_df.columns, fill_value="")
                    prev_add_df = prev_df[prev_df['納品希望日'].astype(str) >= tomorrow]
                    # 合体
                    order_zan_df = pd.concat([order_zan_df, prev_add_df], ignore_index=True)
                    order_zan_df = order_zan_df.drop_duplicates()
                    # シートを一度消して作り直し
                    del wb['注文残']
                    ws_oj = wb.create_sheet('注文残')
                    ws_oj.append(list(order_zan_df.columns))
                    for row in order_zan_df.itertuples(index=False, name=None):
                        ws_oj.append(row)

                # 列幅自動調整
                for ws in wb.worksheets:
                    autofit_columns(ws)
                wb.save(file_path)

                # Drive再アップロード
                media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                drive_service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()
                print("受注残・発注残シート作成＆Drive再アップロード完了！")
            except Exception as e:
                print(f"受注残・発注残作成またはDriveアップロードエラー: {e}")
            return 'OK', 200

        # =====================
        # 発注残・注文残の前日データ移行
        # =====================
        if user_text == '受注残と発注残の前日データ移行':
            try:
                ok = migrate_prev_day_sheets_to_today(csv_folder_id, today, drive_service)
                if not ok:
                    print("前日データ移行に失敗")
                    return 'OK', 200
                print("前日データ移行完了！")
            except Exception as e:
                print(f"前日データ移行エラー: {e}")
            return 'OK', 200

        # --- 通常テキスト（注文等）は既存ハンドラへ ---
        process_text_message(event)

    elif message_type == 'image':
        process_image_message(event)

    elif message_type == 'file':
        file_name = event['message'].get('fileName', '').lower()
        file_id = event['message'].get('fileId')
        print("file_name repr:", repr(file_name))

        # タグ付け表.xlsxの場合はGoogleドライブ受注集計直下にアップロード
        if unicodedata.normalize('NFC', file_name.strip()) == 'タグ付け表.xlsx':
            temp_path = f"/tmp/{file_name}"
            CHANNEL_ACCESS_TOKEN = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
            headers = {"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"}
            url = f"https://api-data.line.me/v2/bot/message/{file_id}/content"
            r = requests.get(url, headers=headers, stream=True)
            with open(temp_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=1024):
                    if chunk:
                        f.write(chunk)
            try:
                root_id = get_or_create_folder('受注集計')
                file_metadata = {'name': file_name, 'parents': [root_id]}
                media = MediaFileUpload(temp_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                drive_service.files().create(
                    body=file_metadata, media_body=media, fields='id', supportsAllDrives=True
                ).execute()
                print("タグ付け表.xlsxをGoogleドライブにアップロードしました")
            except Exception as e:
                print(f"タグ付け表.xlsxのDrive保存エラー: {e}")
            return 'OK', 200

        # 注文書フォーマット.xlsxの場合はGoogleドライブ受注集計直下にアップロード
        elif file_name == '注文書フォーマット.xlsx':
            temp_path = f"/tmp/{file_name}"
            CHANNEL_ACCESS_TOKEN = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
            headers = {"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"}
            url = f"https://api-data.line.me/v2/bot/message/{file_id}/content"
            r = requests.get(url, headers=headers, stream=True)
            with open(temp_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=1024):
                    if chunk:
                        f.write(chunk)
            try:
                root_id = get_or_create_folder('受注集計')
                file_metadata = {'name': file_name, 'parents': [root_id]}
                media = MediaFileUpload(temp_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                drive_service.files().create(
                    body=file_metadata, media_body=media, fields='id', supportsAllDrives=True
                ).execute()
                print("注文書フォーマット.xlsxをGoogleドライブにアップロードしました")
            except Exception as e:
                print(f"注文書フォーマット.xlsxのDrive保存エラー: {e}")
            return 'OK', 200

        # それ以外（PDF等）は既存処理
        elif file_name.endswith('.pdf'):
            process_pdf_message(event)
        # 他のファイル型は必要に応じてハンドラ追加

    return 'OK', 200
