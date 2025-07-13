import pandas as pd
import io
from handlers.file_handler import drive_service
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from config import CSV_FORMAT_PATH, SHARED_DRIVE_ID, ORDER_SUMMARY_FOLDER_ID
import pytz
from datetime import datetime
import unicodedata
import os
from openpyxl import Workbook, load_workbook
import jaconv  # ひらがな→カタカナ正規化用
from openai import OpenAI  # 新しいOpenAIクライアント
from .prompt_templates import normalize_product_name_prompt
import re
import shutil
from handlers.file_handler import get_or_create_folder
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

CSV_HEADERS = pd.read_csv(CSV_FORMAT_PATH, encoding='utf-8').columns.tolist()
JST = pytz.timezone('Asia/Tokyo')

def normalize_product_name_ai(product_name, openai_client):
    # 生成AIでカタカナ統一
    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": normalize_product_name_prompt},
            {"role": "user", "content": product_name}
        ],
        max_tokens=10,
        temperature=0
    )
    return response.choices[0].message.content.strip()

def normalize_size(size):
    # 半角英数字・大文字化
    if pd.isnull(size):
        return ""
    return jaconv.z2h(str(size), kana=False, ascii=True, digit=True).upper().strip()

def normalize_quantity(quantity):
    # 半角数字のみ
    if pd.isnull(quantity):
        return ""
    return jaconv.z2h(str(quantity), kana=False, ascii=False, digit=True).strip()

def normalize_unit_postprocess(unit):
    """
    英字は半角大文字、カタカナは全角化（ひらがな→カタカナも対応）
    """
    if not unit:
        return ""
    unit = str(unit).strip()
    # ひらがな→カタカナ
    unit = jaconv.hira2kata(unit)
    # 全角英字→半角大文字
    unit = jaconv.z2h(unit, kana=False, ascii=True, digit=True)
    unit = re.sub(r'[a-z]', lambda m: m.group(0).upper(), unit)
    # 半角カタカナ→全角カタカナ
    unit = jaconv.h2z(unit, ascii=False, digit=False)
    return unit

def normalize_unit_ai(product_name, unit, quantity, openai_client):
    from .prompt_templates import normalize_unit_prompt
    content = f"商品名: {product_name}\n単位: {unit}\n数量: {quantity}"
    try:
        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": normalize_unit_prompt},
                {"role": "user", "content": content}
            ],
            max_tokens=10,
            temperature=0
        )
        result = response.choices[0].message.content.strip()
        # --- 返答が空、問い合わせ文そのもの、異常系なら元のunitを返す ---
        if (not result or 
            result.startswith("商品名:") or 
            result.startswith("単位:") or 
            "単位" in result or 
            "商品名" in result or 
            len(result) > 10):  # "kg"や"玉"など一般的な単位は2～4文字程度
            return normalize_unit_postprocess(unit)
        # 返答も正規化
        return normalize_unit_postprocess(result)
    except Exception as e:
        print(f"[AI単位正規化エラー] {e} 元の単位({unit})を返却します")
        return normalize_unit_postprocess(unit)

# 必要に応じてOpenAIクライアントをDI
def normalize_row(row, openai_client):
    # 商品名
    product_name = normalize_product_name_ai(row['商品名'], openai_client)
    # サイズ
    size = normalize_size(row['サイズ'])
    # 数量
    quantity = normalize_quantity(row['数量'])
    # 単位のAI正規化
    norm_unit = normalize_unit_ai(product_name, row['単位'], quantity, openai_client)
    # 数量・単位補正
    adj_quantity, adj_unit = adjust_quantity_and_unit(quantity, norm_unit)
    
    # すべてのカラムを残して新しいdictを作る
    new_row = dict(row)  # 元のデータをコピー
    new_row['商品名'] = product_name
    new_row['サイズ'] = size
    new_row['数量'] = adj_quantity
    new_row['単位'] = adj_unit
    return new_row

def normalize_df(df, openai_client):
    # DataFrame全体を正規化
    normalized_rows = [normalize_row(row, openai_client) for _, row in df.iterrows()]
    return pd.DataFrame(normalized_rows)

def adjust_quantity_and_unit(quantity, unit):
    # g系はkgに変換
    if unit in ["g", "グラム", "ｇ"]:
        try:
            return float(quantity) / 1000, "kg"
        except Exception:
            return quantity, unit  # 変換できなければそのまま
    elif unit in ["kg", "キログラム", "ＫＧ"]:
        return quantity, "kg"
    else:
        return quantity, unit

def append_to_xlsx(structured_text, parent_id, openai_client):
    """ 構造化テキストを.xlsxで保存・追記し、Driveに反映（備考コメントなしバージョン） """
    if not structured_text.strip():
        with open("/tmp/failed_structured_text.txt", "w", encoding="utf-8") as f:
            f.write("No structured_text received!\n")
        print("No structured_text received! ログを保存しました。")
        return

    today = datetime.now(JST).strftime('%Y%m%d')
    filename = f'集計結果_{today}.xlsx'
    file_path = f'/tmp/{filename}'

    # 構造化テキスト → DataFrame化
    lines = structured_text.strip().splitlines()
    valid_lines = []
    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue
        if (
            "申し訳ありません" in line_stripped or
            "変換できません" in line_stripped or
            "画像から" in line_stripped or
            "GPT" in line_stripped or
            "ご不明点" in line_stripped or
            line_stripped.count(',') < len(CSV_HEADERS)-1
        ):
            continue
        cols = [c.strip() for c in line_stripped.split(',')]
        if len(cols) == len(CSV_HEADERS):
            valid_lines.append(",".join(cols))

    if not valid_lines:
        print("⚠ 有効な行がありません。全行ログ保存")
        with open(f"/tmp/failed_structured_{today}.txt", "w", encoding="utf-8") as f:
            f.write(structured_text)
        return

    structured_text_cleaned = "\n".join(valid_lines)
    try:
        new_data = pd.read_csv(io.StringIO(structured_text_cleaned), header=None, names=CSV_HEADERS)
    except Exception as e:
        print("CSV parsing error:", e)
        with open(f"/tmp/csv_parse_error_{today}.txt", "w", encoding="utf-8") as f:
            f.write(structured_text)
        return

    now_str = datetime.now(JST).strftime('%Y%m%d%H')
    new_data['時間'] = now_str

    # Drive上の既存ファイル取得＆マージ
    print("\n【デバッグ】Drive全体で見える同名ファイル一覧:")
    all_results = drive_service.files().list(
        q=f"name='{filename}' and trashed = false",
        fields="files(id, name, parents, owners)",
        pageSize=10,
        driveId=SHARED_DRIVE_ID,
        corpora='drive',
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    all_files = all_results.get('files', [])
    for f in all_files:
        print(f"ファイル名: {f['name']}, ファイルID: {f['id']}, 親: {f.get('parents')}, オーナー: {f['owners'][0]['displayName'] if f.get('owners') else '-'}")

    query = f"name = '{filename}' and '{parent_id}' in parents and trashed = false"
    response = drive_service.files().list(
        q=query,
        fields='files(id, name, parents, owners)',
        driveId=SHARED_DRIVE_ID,
        corpora='drive',
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    files = response.get('files', [])
    print(f"【デバッグ】指定親フォルダ {parent_id} で見つかったファイル数: {len(files)}")
    for f in files:
        print(f"【デバッグ】指定親: ファイル名: {f['name']}, ファイルID: {f['id']}, 親: {f.get('parents')}, オーナー: {f['owners'][0]['displayName'] if f.get('owners') else '?'}")

    main_sheet_name = f'集計結果_{today}'
    if files:
        file_id = files[0]['id']
        request = drive_service.files().get_media(
            fileId=file_id,
            supportsAllDrives=True
        )
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        fh.seek(0)
        # 全シート読み込み
        xl = pd.read_excel(fh, sheet_name=None)
        if main_sheet_name in xl:
            existing = xl[main_sheet_name]
            combined = pd.concat([existing, new_data], ignore_index=True)
        else:
            combined = new_data
    else:
        combined = new_data

    # サマリも含めたxlsxで保存＆Drive反映
    from .csv_handler import xlsx_with_summary_update  # 必要に応じて調整
    xlsx_with_summary_update(combined, file_path, openai_client)
    try:
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if files:
            drive_service.files().update(
                fileId=file_id,
                media_body=media,
                supportsAllDrives=True
            ).execute()
        else:
            file_metadata = {'name': filename, 'parents': [parent_id]}
            drive_service.files().create(
                body=file_metadata,
                media_body=media,
                supportsAllDrives=True
            ).execute()
        print(f"Excelファイル作成成功: {file_path}")
    except Exception as e:
        print("Excelファイル作成/アップロードエラー:", e)

def xlsx_with_summary_update(df, xlsx_path, openai_client):
    """
    1シート目: 生データ
    2シート目: 商品名・サイズ・単位・納品希望日ごとの集計サマリ
    ※顧客、発注者、納品場所、時間、社内担当者はサマリ側は空欄に
    """
    # --- 正規化 ---
    normalized_rows = []
    for _, row in df.iterrows():
        # 正規化ロジック（各自のプロジェクトで実装。ここは例）
        product_name = normalize_product_name_ai(row['商品名'], openai_client)
        size = normalize_size(row['サイズ'])
        quantity = normalize_quantity(row['数量'])
        norm_unit = normalize_unit_ai(product_name, row['単位'], quantity, openai_client)
        adj_quantity, adj_unit = adjust_quantity_and_unit(quantity, norm_unit)
        normalized_rows.append({
            "顧客": row.get("顧客", ""),
            "発注者": row.get("発注者", ""),
            "商品名": product_name,
            "サイズ": size,
            "数量": adj_quantity,
            "単位": adj_unit,
            "納品希望日": row.get("納品希望日", ""),
            "納品場所": row.get("納品場所", ""),
            "時間": row.get("時間", ""),
            "社内担当者": row.get("社内担当者", ""),
            "備考": row.get("備考", "")
        })
    df_norm = pd.DataFrame(normalized_rows)

    # --- 数量は必ず数値型で ---
    df_norm['数量'] = pd.to_numeric(df_norm['数量'], errors='coerce').fillna(0)

    # --- 集計キー（納品希望日も追加） ---
    df_norm['集計キー'] = (
        df_norm['商品名'].astype(str) + "_" +
        df_norm['サイズ'].astype(str) + "_" +
        df_norm['単位'].astype(str) + "_" +
        df_norm['納品希望日'].astype(str)
    )

    # --- サマリ生成 ---
    summary = (
        df_norm.groupby('集計キー', as_index=False)
        .agg({
            '顧客': lambda x: "",
            '発注者': lambda x: "",
            '商品名': 'first',
            'サイズ': 'first',
            '数量': 'sum',      # ← 数値として合計！
            '単位': 'first',
            '納品希望日': 'first',
            '納品場所': lambda x: "",
            '時間': lambda x: "",
            '社内担当者': lambda x: "",
            '備考': 'first'
        })
    )
    summary = summary[['顧客', '発注者', '商品名', 'サイズ', '数量', '単位', '納品希望日', '納品場所', '時間', '社内担当者', '備考']]
    summary = summary.sort_values('商品名')

    # --- xlsx出力 ---
    # 既存ファイルがあれば読み込み、なければ新規
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    # 元データシート名（例: ファイル名拡張子抜き）
    raw_sheet_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    if raw_sheet_name in wb.sheetnames:
        del wb[raw_sheet_name]

    ws_raw = wb.create_sheet(raw_sheet_name)
    ws_raw.append(list(df_norm.columns.drop('集計キー')))
    for row in df_norm.drop(columns=['集計キー']).itertuples(index=False, name=None):
        ws_raw.append(row)

    # サマリシート
    if "集計結果サマリ" in wb.sheetnames:
        del wb["集計結果サマリ"]
    ws_summary = wb.create_sheet("集計結果サマリ")
    ws_summary.append(list(summary.columns))
    for row in summary.itertuples(index=False, name=None):
        ws_summary.append(row)

    # 列幅自動調整（autofit_columnsがあれば）
    for ws in wb.worksheets:
        autofit_columns(ws)
    wb.save(xlsx_path)
    print(f"集計結果サマリシート付きで {xlsx_path} を作成しました")

def create_order_list_sheet(xlsx_path, tag_xlsx_path):
    """
    「集計結果サマリ」→「注文リスト」シートを作成
    備考と発注先の間に税率（タグ付け表由来）を追加
    """
    # 必要なヘッダーに「税率」を追加
    order_headers = ["商品名", "サイズ", "数量", "単位", "納品希望日", "備考", "税率", "発注先", "郵便番号", "住所"]

    # サマリ読み込み
    wb = load_workbook(xlsx_path)
    if "集計結果サマリ" not in wb.sheetnames:
        print("集計結果サマリシートがありません")
        return False

    ws = wb["集計結果サマリ"]
    summary_df = pd.DataFrame(ws.values)
    summary_df.columns = summary_df.iloc[0]
    summary_df = summary_df[1:]

    # タグ付け表読み込み
    tag_df = pd.read_excel(tag_xlsx_path, dtype=str).fillna("")

    order_list = []
    for _, row in summary_df.iterrows():
        prod, size = row['商品名'], row['サイズ']

        match = tag_df[(tag_df['商品名'] == prod) & (tag_df['サイズ'] == size)]
        if len(match) == 0:
            match = tag_df[(tag_df['商品名'] == prod) & (tag_df['サイズ'] == "")]
        if len(match) == 0:
            # データがなければ空欄
            supplier, zipcode, address, tax_rate = "", "", "", ""
        else:
            supplier = match.iloc[0].get('発注先', "")
            zipcode = match.iloc[0].get('郵便番号', "")
            address = match.iloc[0].get('住所', "")
            raw_tax = match.iloc[0].get('税率', "")
            if raw_tax != "":
                try:
                    tax_float = float(raw_tax)
                    if tax_float < 1.0:
                        tax_rate = f"{int(round(tax_float * 100))}%"
                    else:
                        tax_rate = f"{int(round(tax_float))}%"
                except Exception:
                    tax_rate = str(raw_tax)
            else:
                tax_rate = ""

        order_list.append([
            prod,
            size,
            row['数量'],
            row['単位'],
            row['納品希望日'],
            row.get('備考', ""),
            tax_rate,        # ★ここに税率
            supplier,
            zipcode,
            address
        ])

    # 既存シート削除
    if "注文リスト" in wb.sheetnames:
        del wb["注文リスト"]

    ws_order = wb.create_sheet("注文リスト")
    ws_order.append(order_headers)
    for r in order_list:
        ws_order.append(list(r))

    # セル値を書き換えるときはMergedCellを絶対に触らない
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue  # 結合セルは触らない

    # 必要に応じて列幅自動調整（関数化している場合はそのまま）
    for ws in wb.worksheets:
        if 'autofit_columns' in globals():
            autofit_columns(ws)

    wb.save(xlsx_path)
    return True

def create_order_sheets(date_id, csv_folder_id, today_str, drive_service):
    """
    「注文リスト」シートから、発注先ごとに「注文書フォーマット.xlsx」をコピー・編集し
    「注文書_YYYYMMDD_連番.xlsx」ファイルをGoogle Drive「注文書」フォルダへアップロードする
    """
    # 1. 注文書フォーマット.xlsx取得
    fmt_query = f"name = '注文書フォーマット.xlsx' and trashed = false"
    fmt_resp = drive_service.files().list(
        q=fmt_query,
        fields='files(id)',
        driveId=SHARED_DRIVE_ID,
        corpora='drive',
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    fmt_files = fmt_resp.get('files', [])
    if not fmt_files:
        print("注文書フォーマット.xlsxが見つかりません")
        return False
    fmt_file_id = fmt_files[0]['id']

    # 2. 注文リストシートのDL
    filename = f'集計結果_{today_str}.xlsx'
    file_path = f"/tmp/{filename}"
    query = f"name = '{filename}' and '{csv_folder_id}' in parents and trashed = false"
    response = drive_service.files().list(
        q=query,
        fields='files(id)',
        driveId=SHARED_DRIVE_ID,
        corpora='drive',
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    files = response.get('files', [])
    if not files:
        print("集計ファイルが見つかりません")
        print("csv_folder_id:", csv_folder_id)
        return False
    excel_file_id = files[0]['id']
    excel_tmp_path = f"/tmp/{filename}"
    request_dl = drive_service.files().get_media(
        fileId=excel_file_id,
        supportsAllDrives=True
    )
    with open(excel_tmp_path, 'wb') as fh:
        downloader = MediaIoBaseDownload(fh, request_dl)
        done = False
        while not done:
            status, done = downloader.next_chunk()

    from openpyxl import load_workbook
    wb = load_workbook(excel_tmp_path)
    if "注文リスト" not in wb.sheetnames:
        print("注文リストシートがありません")
        return False
    df = pd.DataFrame(wb["注文リスト"].values)
    df.columns = df.iloc[0]
    df = df[1:]

    # 3. 注文書フォルダの作成
    from handlers.file_handler import get_or_create_folder
    order_folder_id = get_or_create_folder("注文書", parent_id=date_id)

    # 4. 発注先ごとにグループ
    grouped = df.groupby("発注先")
    count = 1
    for supplier, g in grouped:
        if not supplier or str(supplier).strip() == "":
            continue

        # フォーマットファイルをDL＆コピー
        fmt_tmp_path = f"/tmp/注文書フォーマット.xlsx"
        fmt_dl = drive_service.files().get_media(fileId=fmt_file_id)
        with open(fmt_tmp_path, 'wb') as ffmt:
            downloader = MediaIoBaseDownload(ffmt, fmt_dl)
            done = False
            while not done:
                status, done = downloader.next_chunk()

        dest_name = f"注文書_{today_str}_{count:03d}.xlsx"
        dest_path = f"/tmp/{dest_name}"
        shutil.copy(fmt_tmp_path, dest_path)

        # 注文書記載
        wbo = load_workbook(dest_path)
        ws = wbo.active

        # B6, B7, B8, P4, B15
        ws["B6"] = g.iloc[0]["発注先"]
        ws["B7"] = g.iloc[0]["郵便番号"]
        ws["B8"] = g.iloc[0]["住所"]
        ws["P4"] = pd.Timestamp.today().strftime("%Y/%m/%d")

        # 商品行ループ
        for i, (_, row) in enumerate(g.iterrows()):
            row_idx = 20 + i  # 20~34
            ws[f"B{row_idx}"] = row["商品名"]
            # 消費税欄がなければK列は空欄
            if "消費税" in g.columns:
                ws[f"K{row_idx}"] = "※" if str(row.get("消費税", "")).strip() == "10%" else ""
            ws[f"L{row_idx}"] = row["数量"]
            ws[f"M{row_idx}"] = row["単位"]
            ws[f"F{row_idx}"] = row["サイズ"]
            orig = row["納品希望日"]
            if isinstance(orig, str) and len(orig) == 8 and orig.isdigit():
                dt = datetime.strptime(orig, "%Y%m%d")
                ws[f"H{row_idx}"] = dt.strftime("%Y/%m/%d")
            else:
                ws[f"H{row_idx}"] = orig  # フォーマット外ならそのまま
            if row_idx >= 34:
                break

        wbo.save(dest_path)

        # Driveへアップロード
        file_metadata = {'name': dest_name, 'parents': [order_folder_id]}
        media = MediaFileUpload(dest_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        drive_service.files().create(
            body=file_metadata,
            media_body=media,
            supportsAllDrives=True
        ).execute()
        count += 1

    return True

def create_order_remains_sheet_from_wb(wb):
    try:
        if "注文リスト" not in wb.sheetnames:
            print("注文リストシートがありません")
            return False
        ws_order = wb["注文リスト"]
        import pandas as pd
        df = pd.DataFrame(ws_order.values)
        df.columns = df.iloc[0]
        df = df[1:]

        # 翌日以降のもの
        from datetime import datetime, timedelta
        JST = pytz.timezone('Asia/Tokyo')
        tomorrow = (datetime.now(JST) + timedelta(days=1)).strftime('%Y%m%d')
        order_remains = df[df['納品希望日'].apply(lambda x: str(x).isdigit() and str(x) >= tomorrow)]

        # シート上書き
        if "注文残" in wb.sheetnames:
            del wb["注文残"]
        ws_remains = wb.create_sheet("注文残")
        ws_remains.append(list(df.columns))
        for row in order_remains.itertuples(index=False, name=None):
            ws_remains.append(row)
        print(f"注文残シートを作成: {wb.properties.title or ''}")
        return True
    except Exception as e:
        print(f"注文残シート作成エラー: {e}")
        return False

def migrate_prev_day_sheets_to_today(csv_folder_id, today_str, drive_service):
    """
    前日の「受注残」「注文残」シートを今日の集計エクセルに(前日データ)シートとして移行
    """
    # 日付文字列
    from datetime import datetime, timedelta
    JST = pytz.timezone('Asia/Tokyo')
    today = datetime.now(JST)
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime('%Y%m%d')
    prev_dt = datetime.strptime(today_str, "%Y%m%d") - timedelta(days=1)
    prev_str = prev_dt.strftime("%Y%m%d")

    # ファイル名定義
    today_xlsx = f"集計結果_{today_str}.xlsx"
    prev_xlsx  = f"集計結果_{prev_str}.xlsx"

    # --- 前日ファイルDL
    root_id = get_or_create_folder('受注集計')
    yesterday_folder_id = get_or_create_folder(yesterday_str, parent_id=root_id)
    csv_folder_id_yesterday = get_or_create_folder('集計結果', parent_id=yesterday_folder_id)
    prev_query = f"name = '{prev_xlsx}' and '{csv_folder_id_yesterday}' in parents and trashed = false"
    prev_response = drive_service.files().list(
        q=prev_query,
        fields='files(id)',
        driveId=SHARED_DRIVE_ID,
        corpora='drive',
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    prev_files = prev_response.get('files', [])
    if not prev_files:
        print("前日分の集計結果ファイルがありません")
        return False
    prev_file_id = prev_files[0]['id']
    prev_tmp_path = f"/tmp/{prev_xlsx}"
    request_dl = drive_service.files().get_media(
        fileId=prev_file_id,
        supportsAllDrives=True
    )
    with open(prev_tmp_path, 'wb') as fh:
        downloader = MediaIoBaseDownload(fh, request_dl)
        done = False
        while not done:
            status, done = downloader.next_chunk()
    prev_wb = load_workbook(prev_tmp_path)

    # --- 当日ファイルDL or 新規作成
    today_query = f"name = '{today_xlsx}' and '{csv_folder_id}' in parents and trashed = false"
    today_response = drive_service.files().list(
        q=today_query,
        fields='files(id)',
        driveId=SHARED_DRIVE_ID,
        corpora='drive',
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    today_files = today_response.get('files', [])
    today_tmp_path = f"/tmp/{today_xlsx}"

    if today_files:
        today_file_id = today_files[0]['id']
        request_dl = drive_service.files().get_media(
            fileId=prev_file_id,
            supportsAllDrives=True
        )
        with open(today_tmp_path, 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request_dl)
            done = False
            while not done:
                status, done = downloader.next_chunk()
        today_wb = load_workbook(today_tmp_path)
    else:
        # 新規作成
        today_wb = load_workbook(prev_tmp_path)  # 前日ファイルをコピーして新ファイルにしてもOK
        # あるいは空ブックでも良いが、必要なシート名・ヘッダ構成に注意

    # --- コピー対象シート
    for src_name, dst_name in [("受注残", "受注残(前日データ)"), ("注文残", "注文残(前日データ)")]:
        if src_name in prev_wb.sheetnames:
            # 既存で当日側にあれば削除
            if dst_name in today_wb.sheetnames:
                std = today_wb[dst_name]
                today_wb.remove(std)
            ws_prev = prev_wb[src_name]
            # データをリスト化
            data = list(ws_prev.values)
            # 新規シート作成
            ws_today = today_wb.create_sheet(dst_name)
            for row in data:
                ws_today.append(row)
        else:
            print(f"前日ファイルに{src_name}シートがありません")

    # --- 保存・再アップロード
    for ws in today_wb.worksheets:
        autofit_columns(ws)
    today_wb.save(today_tmp_path)
    media = MediaFileUpload(today_tmp_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if today_files:
        drive_service.files().update(
            fileId=today_file_id,
            media_body=media,
            supportsAllDrives=True
        ).execute()
    else:
        file_metadata = {'name': today_xlsx, 'parents': [csv_folder_id]}
        drive_service.files().create(
            body=file_metadata,
            media_body=media,
            supportsAllDrives=True
        ).execute()
    print("前日データ移行シートを作成・アップロード完了")
    return True

def autofit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = ( max_length + 2 ) * 2  # 余白も考慮
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width